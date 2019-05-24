
# ***************************************
'''
import re
import requests
from bs4 import BeautifulSoup
import urllib.request
import csv
import sys
import os
import zipfile
import logging
import time
import datetime
url1 = 'https://www.sec.gov/Archives/edgar/data/1097718/000135448815004617/uoip_10k.htm'
url0 = 'https://www.sec.gov/Archives/edgar/data/798528/000119312519094529/d657683d10k.htm'
url_1 = 'https://github.com/ragraw26/Edgar-COMPANY-FILINGS-Web-Scrapping-Data-Analysis/blob/master/Data%20Scrapping/Team_5_Part1_Report.pdf'
def get_soup(url):
    try:
        htmlpage = urllib.request.urlopen(url)
        page = BeautifulSoup(htmlpage, "html.parser")
        find_all_tables(page)
    except:
        print('failing in get_soup')
        return None


def find_all_tables(page):
    all_divtables = page.find_all('table')
    find_all_datatables(page, all_divtables)
    return 0

def foldername(page):
    title = page.find('filename').contents[0]
    if ".htm" in title:
        foldername = title.split(".htm")
        return foldername[0]

def zip_dir(path_dir, path_file_zip=''):
    if not path_file_zip:
        path_file_zip = os.path.join(os.path.dirname(path_dir), os.path.basename(path_dir) + '.zip')
    with zipfile.ZipFile(path_file_zip, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for root, dirs, files in os.walk(path_dir):
            for file_or_dir in files + dirs:
                zip_file.write(os.path.join(root, file_or_dir),os.path.relpath(os.path.join(root, file_or_dir),os.path.join(path_dir, os.path.pardir)))


def assure_path_exists(path):
    if not os.path.exists(path):
        os.makedirs(path)


def checktag(param):
    setflag = "false"
    datatabletags = ["background", "bgcolor", "background-color"]
    for x in datatabletags:
        if x in param:
            setflag = "true"
    return setflag

def checkheadertag(param):
    setflag="false"
    datatabletags=["center","bold"]
    for x in datatabletags:
        if x in param:
            setflag="true"
    return setflag


def printtable(table):
    logging.debug('In a function : printtable')
    printtable = []
    printtrs = table.find_all('tr')
    for tr in printtrs:
        data=[]
        pdata=[]
        printtds=tr.find_all('td')
        for elem in printtds:
            x=elem.text
            x=re.sub(r"['()]","",str(x))
            x=re.sub(r"[$]"," ",str(x))
            if(len(x)>1):
                x=re.sub(r"[—]","",str(x))
                pdata.append(x)
        data=([elem.encode('utf-8') for elem in pdata])
        printtable.append([elem.decode('utf-8').strip() for elem in data])
    return printtable


def find_all_datatables(page, all_divtables):
    count = 0
    allheaders=[]
    for table in all_divtables:
        bluetables = []
        trs = table.find_all('tr')
        for tr in trs:
            # global flagtr
            if checktag(str(tr.get('style'))) == "true" or checktag(str(tr)) == "true":
                bluetables = printtable(tr.find_parent('table'))
                break
            else:
                tds = tr.find_all('td')
                for td in tds:
                    if checktag(str(td.get('style'))) == "true" or checktag(str(td)) == "true":
                        bluetables = printtable(td.find_parent('table'))
                        break
            if not len(bluetables) == 0:
                break
        if not len(bluetables) == 0:
            count += 1
            ptag=table.find_previous('p')
            while ptag is not None and checkheadertag(ptag.get('style'))=="false" and len(ptag.text)<=1:
                ptag=ptag.find_previous('p')
                if checkheadertag(ptag.get('style'))=="true" and len(ptag.text)>=2:
                    global name
                    name=re.sub(r"[^A-Za-z0-9]+","",ptag.text)
                    if name in allheaders:
                        hrcount+=1
                        hrname=name+"_"+str(hrcount)
                        allheaders.append(hrname)
                    else:
                        hrname=name
                        allheaders.append(hrname)
                        break
            folder_name = foldername(page)
            path = str(os.getcwd()) + "/" + folder_name
            assure_path_exists(path)
            if(len(allheaders)==0):
                filename=folder_name+"-"+str(count)
            else:
                filename=allheaders.pop()
            csvname=filename+".csv"
            csvpath = path + "/" + csvname
            with open(csvpath, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)
                writer.writerows(bluetables)
            zip_dir(path)


url2 = 'file:///Users/bmandyam/Downloads/uoip_10k.html'
get_soup(url2)

'''

# **********************************************************






# import requests
from bs4 import BeautifulSoup
import urllib.request
import re
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from nltk import tokenize

bookname = 'test.xlsx'

def checktag(param):
    setflag = False
    datatabletags = ["background", "bgcolor", "background-color"]
    for x in datatabletags:
        if x in param:
            setflag = True
    return setflag

# def checktag(param):
#     setflag = False
#     if '#cceeff' in param:
#         setflag = True
#     return setflag

def scrape_each_table(table):
    rows = table.find_all('tr')
    table_data = []
    for tr in rows:
        row_data=[]
        orgRdata=[]
        cells = tr.find_all('td')
        for cell in cells:
            cell_data = cell.text
            # cell_data =re.sub(r"['()]","",str(cell_data))
            cell_data = re.sub(r"[$]","",str(cell_data))
            # repl_list = [r"['()]", r"[$]"]
            # cell_data=re.sub("|".join(repl_list),"",str(cell_data))
            if(len(cell_data)>0):
                # cell_data=re.sub(r"[—]","",str(cell_data))
                row_data.append(cell_data.encode('utf-8'))
                
        table_data.append([x.decode('utf-8').strip() for x in row_data])
    return table_data
        
def save_to_excel(df,i):
    global bookname
    # print(bookname)
    book = openpyxl.load_workbook(bookname)
    book.create_sheet(str(i))
    sheet = book[str(i)]

    for r in dataframe_to_rows(df, index=False, header=False):
        sheet.append(r)
    book.save(bookname)

def create_xlsx_workbook(url):
    global bookname
    bookname = url.split("/")[-1].split('.')[0] + '.xlsx'
    wb = openpyxl.Workbook()
    wb.save(bookname)

def extract_heading_from_text(heading):
    heading_identifiers = ["below", "follows", "following"]
    for x in heading_identifiers:
        if x in heading.lower():
            sentences_list = tokenize.sent_tokenize(heading)
            for sentence in sentences_list:
                if x in sentence.lower():
                    return sentence

    return ''


def find_heading(table):
    # print('inside find heading')
    i=0
    table_flag = 0
    bold_heading_flag = False
    normal_heading_flag = False
    heading  = ''
    div_with_text = table
    while True:
        # print('div',i)
        # print('------',heading)
        i+=1
        div_with_text = div_with_text.find_previous('div')
        if len(div_with_text.find_all('table')) > 0:
            # print('tablecheck fun')
            if table_flag < 2:
                table_flag += 1
                if not heading:
                    continue
                else :
                    break
            break
        else :
            # print(div_with_text)
            # print(div_with_text.text)
            if not div_with_text.text.strip():
                continue
            else :
                heading_txt  = extract_heading_from_text(div_with_text.text)
                
                if not heading_txt:
                    if not normal_heading_flag:
                        heading = div_with_text.text
                        normal_heading_flag = True

                    if 'font-weight' in str(div_with_text).lower():
                        # print('inside bold')
                        # print(bold_heading_flag)
                        

                        if 'consolidated' in div_with_text.text.lower() or 'statements' in div_with_text.text.lower():
                            heading = div_with_text.text
                            break
                        elif bold_heading_flag:
                            break
                        # print('@@@@@@@@@@@@@@@@@@@@@@@@@',div_with_text.text)
                        heading = div_with_text.text
                        bold_heading_flag = True
                        continue
                    
                else :
                    heading = heading_txt
                    break
                

    return heading
        


def arrange_table():
    pass



def main(url):
    create_xlsx_workbook(url)
    htmlpage = urllib.request.urlopen(url)
    page = BeautifulSoup(htmlpage, "html.parser")
    all_divtables = page.find_all('table')


    for i,table in enumerate(all_divtables[26:27]):
        # print(i)
        scraped_tables = []
        rows = table.find_all('tr')
        for tr in rows:    
            if checktag(str(tr)):
                scraped_tables = scrape_each_table(tr.find_parent('table'))
                break
            else:
                tds = tr.find_all('td')
                for td in tds:
                    if checktag(str(td)):
                        scraped_tables = scrape_each_table(td.find_parent('table'))
                        break
            if not len(scraped_tables) == 0:
                break

        if not len(scraped_tables) == 0:
            # print(scraped_tables)
            # print('find heading after this')
            heading = find_heading(table)
            # print('inside main heading***********:',heading)
            # print('\n\nhere\n\n',table.find_previous('div').find_previous('div').find_previous('div'))
            if not heading:
            #     save_to_excel(pd.DataFrame(scraped_tables),i)
                print(scraped_tables)
            else :
                scraped_tables.insert(0,[])
                scraped_tables.insert(1,[heading])
                scraped_tables.insert(2,[])
                # print(scraped_tables)
                # arrange_table(scraped_tables)

                save_to_excel(pd.DataFrame(scraped_tables),i)
            
if __name__== "__main__":
    # url = 'file:///Users/bmandyam/Downloads/uoip_10k.html'
    url = 'file:///Users/bmandyam/Downloads/aapl_10k.html'
    main(url)